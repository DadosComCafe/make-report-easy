import logging

import ipdb
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.DEBUG)


def get_sheet(path: str) -> Workbook:
    workbook = load_workbook(path)
    return workbook


def generate_new_sheet(path: str) -> str:
    workbook = get_sheet(path=path)
    workbook.create_sheet("NumericMetrics")
    new_path = f"{path.replace('.xlsx', '_numeric_report.xlsx')}"
    workbook.save(new_path)
    return new_path


def write_metrics(path: str):
    numeric_path = path
    report_path = path.replace(".xlsx", "_numeric_report.xlsx")
    wb_original = load_workbook(numeric_path)

    wb_original.save(report_path)
    wb_new = load_workbook(report_path)

    wb_new.create_sheet("NumericReport")
    wb_new.save(report_path)

    sheets: list = wb_original.sheetnames

    for sheet in sheets:
        if sheet == "NumericColumns":
            current_sheet = sheet

    col_names = [cel.value for cel in wb_original[current_sheet][1]]
    wb_analise = wb_new["NumericReport"]

    for col_index, name in enumerate(col_names, start=1):
        col_letter = get_column_letter(col_index)
        logging.info("Gerando a Somatória")

        # definindo largura da coluna
        wb_analise.column_dimensions[f"{col_letter}"].width = 55

        # definindo fonte da linha
        wb_analise[f"{col_letter}1"].font = Font(
            name="Calibri", size=16, bold=True, color="0000FF"
        )

        # definindo a altura da linha
        wb_analise.row_dimensions[1].height = 30
        wb_analise[f"{col_letter}1"] = f"Somatória dos valores de: {name}"

        wb_analise[f"{col_letter}2"] = f"=SUM({sheet}!{col_letter}:{col_letter})"
        logging.info("Gerando a média")
        # definindo fonte da linha
        wb_analise[f"{col_letter}5"].font = Font(
            name="Calibri", size=16, bold=True, color="0000FF"
        )

        wb_analise.row_dimensions[5].height = 30
        wb_analise[f"{col_letter}5"] = f"Média dos valores de: {name}"

        wb_analise[f"{col_letter}6"] = f"=AVERAGE({sheet}!{col_letter}:{col_letter})"
        logging.info("Gerando o máximo")

        # definindo fonte da linha
        wb_analise[f"{col_letter}9"].font = Font(
            name="Calibri", size=16, bold=True, color="0000FF"
        )
        wb_analise.row_dimensions[9].height = 30
        wb_analise[f"{col_letter}9"] = f"Máximo dos valores de: {name}"

        wb_analise[f"{col_letter}10"] = f"=MAX({sheet}!{col_letter}:{col_letter})"

        logging.info("Gerando o minimo")
        # definindo fonte da linha
        wb_analise[f"{col_letter}13"].font = Font(
            name="Calibri", size=16, bold=True, color="0000FF"
        )
        wb_analise.row_dimensions[13].height = 30
        wb_analise[f"{col_letter}13"] = f"Mínimo dos valores de: {name}"

        wb_analise[f"{col_letter}14"] = f"=MIN({sheet}!{col_letter}:{col_letter})"

    wb_new.save(report_path)
    logging.info("Relatório gerado com sucesso.")


if __name__ == "__main__":
    path = "assets/file_sample_numeric_only.xlsx"
    # new_path = generate_new_sheet(path=path)
    report = write_metrics(path=path)
