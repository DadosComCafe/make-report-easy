import logging
from random import randint
from typing import List

import ipdb
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

logging.basicConfig(level=logging.DEBUG)


def is_list_string(input_list: list) -> bool:
    """Checa se todos os elementos da lista `input_list` são tipo texto (str).

    Args:
        input_list (list): A lista que será checada.

    Returns:
        bool: True se todos os elementos forem tipo texto, False se não forem.
    """

    return all(isinstance(item, str) for item in input_list)


def is_list_numeric(input_list: list) -> bool:
    """Checa se todos os elementos da lista `input_list` são numéricos (integers ou floats).

    Args:
        input_list (list): A lista que será checada.

    Returns:
        bool: True se todos os elementos forem numéricos, False se não forem.
    """

    return all(isinstance(item, (int, float)) for item in input_list)


def get_numeric_columns(path: str) -> List[int]:
    """Encontra todas as colunas que representam apenas valores numéricos.
    Para definir se a coluna é numérica, usa-se o primeiro, o último, e um valor de posição
    aleatória dentre estes intervalos, a fim de garantir que a coluna seja de fato numérica sem
    necessariamente processar a coluna inteira.

    Args:
        path (str): O caminho do arquivo xlsx que será examinado

    Returns:
        List[int]: Uma lista com as posições de colunas numéricas do arquivo xlsx analisado.
    """

    dict_of_lists = {}
    list_of_ids = []
    worksheet = load_workbook(path)
    worksheet = worksheet.active
    col_idx = 1

    logging.info("Coletando valores das colunas...")
    for i in range(col_idx, worksheet.max_column + 1):
        min_index = worksheet.min_row + 1
        max_index = worksheet.max_row
        random = randint(min_index + 2, max_index + 1) - 1
        column_values = [
            worksheet.cell(row=row_idx, column=i).value
            for row_idx in [
                worksheet.min_row + 1,
                worksheet.min_row + random,
                worksheet.max_row,
            ]
        ]
        dict_of_lists.update({i: column_values})

    logging.info("Verificando o tipo dos valores...")
    logging.info(f"Dicio: {dict_of_lists}")

    for key, value in dict_of_lists.items():
        if is_list_numeric(value):
            list_of_ids.append(key)
    return list_of_ids


def get_string_columns(path: str) -> List[int]:
    """Encontra todas as colunas que representam apenas valores tipo texto (str).
    Para definir se a coluna é tipo texto, usa-se o primeiro, o último, e um valor de posição
    aleatória dentre estes intervalos, a fim de garantir que a coluna seja de fato tipo texto sem
    necessariamente processar a coluna inteira.

    Args:
        path (str): O caminho do arquivo xlsx que será examinado

    Returns:
        List[int]: Uma lista com as posições de colunas tipo texto do arquivo xlsx analisado.
    """
    dict_of_lists = {}
    list_of_ids = []
    worksheet = load_workbook(path)
    worksheet = worksheet.active
    col_idx = 1

    logging.info("Coletando valores das colunas...")
    for i in range(col_idx, worksheet.max_column + 1):
        min_index = worksheet.min_row + 1
        max_index = worksheet.max_row
        random = randint(min_index + 2, max_index + 1) - 1
        column_values = [
            worksheet.cell(row=row_idx, column=i).value
            for row_idx in [
                worksheet.min_row + 1,
                worksheet.min_row + random,
                worksheet.max_row,
            ]
        ]
        dict_of_lists.update({i: column_values})

        logging.info("Verificando o tipo dos valores...")
        logging.info(f"Dicio: {dict_of_lists}")

        for key, value in dict_of_lists.items():
            if is_list_string(value):
                list_of_ids.append(key)
    return set(list_of_ids)


def create_only_numeric_sheet(path: str) -> None:
    # TODO: Melhorar o código desta função
    # criar uma classe
    # essa classe tem as propriedades workbook e worksheet
    # métodos da classe fazem todo o necessário para exportar a
    # planilha numérica e de string
    # assim, uma próxima classe produz a planilha que é o relatório

    workbook = load_workbook(path)
    worksheet = workbook.active

    new_workbook = Workbook()
    new_worksheet = new_workbook.active
    new_worksheet.title = "NumericColumns"
    numeric_columns = get_numeric_columns(path=path)

    for new_col_index, col_id in enumerate(numeric_columns, start=1):
        col_letter = get_column_letter(col_id)
        for row_index, cell in enumerate(worksheet[col_letter], start=1):
            new_worksheet.cell(row=row_index, column=new_col_index, value=cell.value)

    logging.info(f"Criado um sheet com {len(numeric_columns)} colunas numéricas.")

    output_path = path.split(".xlsx")[0] + "_numeric_only.xlsx"
    new_workbook.save(output_path)
    logging.info("Exportado com sucesso!")


def create_only_string_sheet(path: str) -> None:
    workbook = load_workbook(path)
    worksheet = workbook.active

    new_workbook = Workbook()
    new_worksheet = new_workbook.active
    new_worksheet.title = "NumericColumns"
    text_columns = get_string_columns(path=path)

    for new_col_index, col_id in enumerate(text_columns, start=1):
        col_letter = get_column_letter(col_id)
        for row_index, cell in enumerate(worksheet[col_letter], start=1):
            new_worksheet.cell(row=row_index, column=new_col_index, value=cell.value)

    logging.info(f"Criado um sheet com {len(text_columns)} colunas de texto.")

    output_path = path.split(".xlsx")[0] + "_text_only.xlsx"
    new_workbook.save(output_path)
    logging.info("Exportado com sucesso!")
    # TODO: Continuar


if __name__ == "__main__":
    path = "assets/file_sample.xlsx"
    create_only_string_sheet(path=path)
    # create_only_numeric_sheet(path=path)

    # path = "assets/file_sample.xlsx"
    # logging.info("Encontrando colunas numéricas...")
    # list_numeric_columns = get_numeric_columns(path=path)
    # logging.info(f"Colunas numéricas: {list_numeric_columns}")

    # logging.info("..." * 20)

    # logging.info("Encontrando colunas string...")
    # list_string_columns = get_string_columns(path=path)
    # logging.info(f"Colunas de texto: {list_string_columns}")

    # logging.info("Iniciando geração de relatório...")

    # create_only_numeric_sheet(path=path)

    # logging.info(f"Relatório pode ser acessado em {path}.")
