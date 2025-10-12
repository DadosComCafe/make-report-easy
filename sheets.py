import logging
from random import randint
from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from utils import is_list_numeric, is_list_string


class NumericSheet:
    def __init__(self, path: str):
        self.path = path
        self.workbook = load_workbook(self.path)
        self.worksheet = self.workbook.active

        self.new_workbook = Workbook()
        self.new_worksheet = self.new_workbook.active
        self.new_worksheet.title = "NumericColumns"

    def get_numeric_columns(self) -> List:
        dict_of_lists = {}
        list_of_ids = []
        col_idx = 1

        logging.info("Coletando valores das colunas...")
        for i in range(col_idx, self.worksheet.max_column + 1):
            min_index = self.worksheet.min_row + 1
            max_index = self.worksheet.max_row
            random = randint(min_index + 2, max_index + 1) - 1
            column_values = [
                self.worksheet.cell(row=row_idx, column=i).value
                for row_idx in [
                    self.worksheet.min_row + 1,
                    self.worksheet.min_row + random,
                    self.worksheet.max_row,
                ]
            ]
            dict_of_lists.update({i: column_values})

        logging.info("Verificando o tipo dos valores...")
        logging.info(f"Dicio: {dict_of_lists}")

        for key, value in dict_of_lists.items():
            if is_list_numeric(value):
                list_of_ids.append(key)
        return list_of_ids

    def create_only_numeric_sheet(self) -> None:
        self.numeric_columns = self.get_numeric_columns()

        for new_col_index, col_id in enumerate(self.numeric_columns, start=1):
            col_letter = get_column_letter(col_id)
            for row_index, cell in enumerate(self.worksheet[col_letter], start=1):
                self.new_worksheet.cell(
                    row=row_index, column=new_col_index, value=cell.value
                )

        logging.info(
            f"Criado um sheet com {len(self.numeric_columns)} colunas numéricas."
        )

        self.output_path = path.split(".xlsx")[0] + "_numeric_only.xlsx"
        self.new_workbook.save(self.output_path)
        logging.info("Exportado com sucesso!")


class StringSheet:
    def __init__(self, path: str):
        self.path = path
        self.workbook = load_workbook(self.path)
        self.worksheet = self.workbook.active

        self.new_workbook = Workbook()
        self.new_worksheet = self.new_workbook.active
        self.new_worksheet.title = "StringColumns"

    def get_string_columns(self) -> List:
        dict_of_lists = {}
        list_of_ids = []
        col_idx = 1

        logging.info("Coletando valores das colunas...")
        for i in range(col_idx, self.worksheet.max_column + 1):
            min_index = self.worksheet.min_row + 1
            max_index = self.worksheet.max_row
            random = randint(min_index + 2, max_index + 1) - 1
            column_values = [
                self.worksheet.cell(row=row_idx, column=i).value
                for row_idx in [
                    self.worksheet.min_row + 1,
                    self.worksheet.min_row + random,
                    self.worksheet.max_row,
                ]
            ]
            dict_of_lists.update({i: column_values})

        logging.info("Verificando o tipo dos valores...")
        logging.info(f"Dicio: {dict_of_lists}")

        for key, value in dict_of_lists.items():
            if is_list_string(value):
                list_of_ids.append(key)
        return list_of_ids

    def create_only_string_sheet(self) -> None:
        self.string_columns = self.get_string_columns()

        for new_col_index, col_id in enumerate(self.string_columns, start=1):
            col_letter = get_column_letter(col_id)
            for row_index, cell in enumerate(self.worksheet[col_letter], start=1):
                self.new_worksheet.cell(
                    row=row_index, column=new_col_index, value=cell.value
                )

        logging.info(
            f"Criado um sheet com {len(self.string_columns)} colunas tipo texto."
        )

        self.output_path = path.split(".xlsx")[0] + "_string_only.xlsx"
        self.new_workbook.save(self.output_path)
        logging.info("Exportado com sucesso!")


if __name__ == "__main__":
    path = "assets/file_sample.xlsx"
    numeric_sheet_obj = NumericSheet(path=path)
    print(numeric_sheet_obj.get_numeric_columns())
    numeric_sheet_obj.create_only_numeric_sheet()

    string_sheet_obj = StringSheet(path=path)
    print(string_sheet_obj.get_string_columns())
    string_sheet_obj.create_only_string_sheet()
