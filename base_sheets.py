from typing import List

from openpyxl import Workbook, load_workbook

from abstract_sheets import AbstractSheet


class BaseSheet(AbstractSheet):
    def __init__(self, path: str):
        self.path = path
        self.workbook = load_workbook(self.path)
        self.worksheet = self.workbook.active

        self.new_workbook = Workbook()
        self.new_worksheet = self.new_workbook.active
        self.new_worksheet.title = "UniqueTypeColumns"  # para cada Class Sheet, usar o super para sobescrever esta propriedade

    def get_uniquetype_columns(self) -> List:
        raise NotImplementedError("Method not implemented yet!")

    def create_unique_type_sheet(self) -> None:
        raise NotImplementedError("Method not implemented yet!")


if __name__ == "__main__":
    path = "assets/file_sample.xlsx"
    base_sheet_obj = BaseSheet(path=path)
