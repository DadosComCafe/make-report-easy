import logging
from random import randint
from typing import List

from abstract_sheets import AbstractSheet
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from utils import is_list_numeric, is_list_string


class BaseSheet(AbstractSheet):
    def __init__(self, path: str):
        self.path = path
        self.workbook = load_workbook(self.path)
        self.worksheet = self.workbook.active

        self.new_workbook = Workbook()
        self.new_worksheet = self.new_workbook.active
        self.new_worksheet.title = "UniqueTypeColumns"

    def get_uniquetype_columns(self, unique_type: str) -> List:
        dict_of_lists = {}
        list_of_ids = []
        col_idx = 1

        logging.info("Coletando valores das colunas...")
        for i in range(col_idx, self.worksheet.max_column + 1):
            min_index = self.worksheet.min_row + 1
            max_index = self.worksheet.max_row
            random = randint(min_index + 2, max_index + 1) - 1
            column_values = [
                self.worksheet.cell(row=row_idx, column=i).value
                for row_idx in [
                    self.worksheet.min_row + 1,
                    self.worksheet.min_row + random,
                    self.worksheet.max_row,
                ]
            ]
            dict_of_lists.update({i: column_values})

        logging.info("Verificando o tipo dos valores...")
        logging.info(f"Dicio: {dict_of_lists}")

        if unique_type == "numeric":
            for key, value in dict_of_lists.items():
                if is_list_numeric(value):
                    list_of_ids.append(key)
            return list_of_ids

        if unique_type == "text":
            for key, value in dict_of_lists.items():
                if is_list_string(value):
                    list_of_ids.append(key)
            return list_of_ids

    def create_unique_type_sheet(self, unique_type: str) -> None:
        if unique_type == "text":
            self.unique_columns = self.get_uniquetype_columns(unique_type="text")
        else:
            self.unique_columns = self.get_uniquetype_columns(unique_type="numeric")

        for new_col_index, col_id in enumerate(self.unique_columns, start=1):
            col_letter = get_column_letter(col_id)
            for row_index, cell in enumerate(self.worksheet[col_letter], start=1):
                self.new_worksheet.cell(
                    row=row_index, column=new_col_index, value=cell.value
                )

        logging.info(
            f"Criado um sheet com {len(self.unique_columns)} colunas tipo {unique_type}."
        )

        if unique_type == "text":
            self.output_path = self.path.split(".xlsx")[0] + "_string_only.xlsx"
        else:
            self.output_path = self.path.split(".xlsx")[0] + "_numeric_only.xlsx"
        self.new_workbook.save(self.output_path)
        logging.info("Exportado com sucesso!")


if __name__ == "__main__":
    path = "assets/file_sample.xlsx"
    base_sheet_obj = BaseSheet(path=path)
